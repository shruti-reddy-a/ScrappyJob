import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def generate_excel(jobs: List[Dict[str, Any]], filename: str) -> str:
    df = pd.DataFrame(jobs)
    if df.empty:
        df = pd.DataFrame(
            columns=["Source ATS", "Job Title", "Company", "Location", "Application Link", "Posted Date", "Posting Time", "Snippet/Notes"]
        )
    if "_iso_dt" in df.columns:
        df = df.drop(columns=["_iso_dt"])
    
    writer = pd.ExcelWriter(filename, engine="openpyxl")
    df.to_excel(writer, index=False, sheet_name="Job Postings")
    worksheet = writer.sheets["Job Postings"]
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    link_col_idx = 5
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=link_col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.value = "Apply Here"
            cell.font = Font(color="0563C1", underline="single")
    for i, col in enumerate(df.columns):
        column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(i + 1)].width = min(column_len, 50)
    writer.close()
    return filename
