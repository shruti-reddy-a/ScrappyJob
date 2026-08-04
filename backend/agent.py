import os
import json
import uuid
import smtplib
import time
from datetime import datetime
from email.message import EmailMessage
from typing import List, Dict, Any

from dotenv import load_dotenv
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1.base_query import FieldFilter

from firecrawl import FirecrawlApp

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

load_dotenv()


class JobLogger:
    def __init__(self):
        self.logs = []
    
    def log(self, msg):
        print(msg)
        self.logs.append({
            "timestamp": datetime.now().isoformat(),
            "message": str(msg)
        })


# ==========================================
# Configuration & Initialization
# ==========================================
def initialize_firebase():
    if not firebase_admin._apps:
        # Check for service account path or raw JSON string
        sa_path = os.getenv("FIREBASE_SERVICE_ACCOUNT_PATH")
        sa_json_str = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON")

        if sa_path and os.path.exists(sa_path):
            cred = credentials.Certificate(sa_path)
        elif sa_json_str:
            cred_dict = json.loads(sa_json_str)
            cred = credentials.Certificate(cred_dict)
        else:
            # Fallback for default application credentials
            cred = credentials.ApplicationDefault()

        firebase_admin.initialize_app(cred)
    return firestore.client()


def get_drive_service():
    scopes = ["https://www.googleapis.com/auth/drive"]
    sa_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
    sa_json_str = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")

    if sa_path and os.path.exists(sa_path):
        creds = service_account.Credentials.from_service_account_file(
            sa_path, scopes=scopes
        )
    elif sa_json_str:
        cred_dict = json.loads(sa_json_str)
        creds = service_account.Credentials.from_service_account_info(
            cred_dict, scopes=scopes
        )
    else:
        raise ValueError("Google Service Account credentials not provided.")

    return build("drive", "v3", credentials=creds)


# ==========================================
# Scraping Logic
# ==========================================
def scrape_jobs(
    config: Dict[str, Any], firecrawl: FirecrawlApp, db, user_id: str, job_id: str, logger: JobLogger
) -> List[Dict[str, Any]]:
    # Extract config
    job_titles = config.get("job_titles", [])
    locations = config.get("locations", [])
    target_ats = config.get("target_ats", [])

    all_jobs = []
    seen_urls = set()

    # Map friendly names to domains
    ats_domains = {
        "Greenhouse": "boards.greenhouse.io",
        "Ashby": "jobs.ashbyhq.com",
        "Lever": "jobs.lever.co",
        "Workday": "myworkdayjobs.com",
        "iCIMS": "icims.com",
        "BambooHR": "bamboohr.com",
        "Workable": "apply.workable.com",
        "LinkedIn": "linkedin.com/jobs",
        "Indeed": "indeed.com",
    }

    # Construct search prompt based on config
    title_str = ", ".join(job_titles)
    loc_str = ", ".join(locations)

    schema = {
        "type": "object",
        "properties": {
            "jobs": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "job_title": {"type": "string"},
                        "company": {"type": "string"},
                        "location": {"type": "string"},
                        "application_link": {"type": "string"},
                        "posted_date": {"type": "string"},
                        "snippet": {"type": "string"},
                    },
                    "required": ["job_title", "company", "application_link"],
                },
            }
        },
        "required": ["jobs"],
    }

    total_ats = len(target_ats)
    ats_completed = 0

    progress_ref = db.collection("run_progress").document(job_id)
    progress_ref.set(
        {
            "status": "IN_PROGRESS",
            "current_ats": "Starting...",
            "jobs_found_so_far": 0,
            "ats_completed": 0,
            "total_ats": total_ats,
            "updated_at": firestore.SERVER_TIMESTAMP,
        }
    )

    for ats in target_ats:
        # Check for STOP command
        doc = progress_ref.get()
        if doc.exists and doc.to_dict().get("command") == "STOP":
            logger.log("Received STOP command. Halting execution.")
            progress_ref.update(
                {
                    "status": "CANCELLED",
                    "current_ats": "Cancelled by user",
                    "command": firestore.DELETE_FIELD,
                    "updated_at": firestore.SERVER_TIMESTAMP,
                }
            )
            return all_jobs

        if ats not in ats_domains:
            continue

        progress_ref.update(
            {"current_ats": ats, "updated_at": firestore.SERVER_TIMESTAMP}
        )

        logger.log(f"\n--- Scraping {ats} ---")
        domain = ats_domains[ats]
        prompt = f"Find recently posted jobs matching titles: {title_str} in locations: {loc_str}."

        max_retries = 3
        for attempt in range(max_retries):
            try:
                logger.log(f"Scraping {ats} (Attempt {attempt+1}/{max_retries})...")
                response = firecrawl.extract(
                    urls=[f"https://{domain}"],
                    prompt=f"{prompt} Search specifically for: {' OR '.join(job_titles)}",
                    schema=schema,
                    enable_web_search=True,
                )

                if response and response.success:
                    extracted_data = (
                        response.data.get("jobs", [])
                        if isinstance(response.data, dict)
                        else []
                    )
                    for job in extracted_data:
                        url = job.get("application_link", "")
                        if url and url not in seen_urls:
                            seen_urls.add(url)
                            all_jobs.append(
                                {
                                    "Source ATS": ats,
                                    "Job Title": job.get("job_title", "N/A"),
                                    "Company": job.get("company", "N/A"),
                                    "Location": job.get("location", "N/A"),
                                    "Application Link": url,
                                    "Posted Date": job.get("posted_date", "N/A"),
                                    "Snippet/Notes": job.get("snippet", ""),
                                }
                            )
                break  # Success, break out of retry loop
            except Exception as e:
                error_str = str(e)
                logger.log(f"Error scraping {ats}: {error_str}")
                if "Rate Limit" in error_str or "429" in error_str:
                    if attempt < max_retries - 1:
                        logger.log("Rate limit hit. Waiting 60 seconds before retrying...")
                        progress_ref.update(
                            {
                                "current_ats": "Rate limit hit. Retrying in 60s...",
                                "updated_at": firestore.SERVER_TIMESTAMP,
                            }
                        )
                        # Check for stop during wait
                        should_stop = False
                        for _ in range(12):
                            time.sleep(5)
                            doc = progress_ref.get()
                            if doc.exists and doc.to_dict().get("command") == "STOP":
                                should_stop = True
                                break
                        if should_stop:
                            progress_ref.update(
                                {
                                    "status": "CANCELLED",
                                    "current_ats": "Cancelled by user",
                                    "command": firestore.DELETE_FIELD,
                                    "updated_at": firestore.SERVER_TIMESTAMP,
                                }
                            )
                            return all_jobs
                    else:
                        logger.log(f"Max retries reached for {ats}.")
                else:
                    break  # Not a rate limit error, don't retry

        ats_completed += 1
        progress_ref.update(
            {
                "jobs_found_so_far": len(all_jobs),
                "ats_completed": ats_completed,
                "updated_at": firestore.SERVER_TIMESTAMP,
            }
        )

        # Respect Firecrawl free tier rate limit of 2 requests/min by sleeping for 30s between requests
        if ats != target_ats[-1]:
            logger.log("Sleeping for 30 seconds to respect rate limits...")
            progress_ref.update(
                {
                    "current_ats": "Waiting for rate limit (30s)...",
                    "updated_at": firestore.SERVER_TIMESTAMP,
                }
            )
            # Sleep in chunks to allow responsive cancellation
            for _ in range(6):
                time.sleep(5)
                doc = progress_ref.get()
                if doc.exists and doc.to_dict().get("command") == "STOP":
                    logger.log("Received STOP command during rate limit wait. Halting execution.")
                    progress_ref.update(
                        {
                            "status": "CANCELLED",
                            "current_ats": "Cancelled by user",
                            "command": firestore.DELETE_FIELD,
                            "updated_at": firestore.SERVER_TIMESTAMP,
                        }
                    )
                    return all_jobs

    progress_ref.update(
        {
            "status": "COMPLETED",
            "current_ats": "Finished",
            "updated_at": firestore.SERVER_TIMESTAMP,
        }
    )

    return all_jobs


# ==========================================
# Excel Generation
# ==========================================
def generate_excel(jobs: List[Dict[str, Any]], filename: str) -> str:
    df = pd.DataFrame(jobs)
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "Source ATS",
                "Job Title",
                "Company",
                "Location",
                "Application Link",
                "Posted Date",
                "Snippet/Notes",
            ]  # type: ignore
        )

    writer = pd.ExcelWriter(filename, engine="openpyxl")
    df.to_excel(writer, index=False, sheet_name="Job Postings")

    worksheet = writer.sheets["Job Postings"]

    # Header Styling
    header_fill = PatternFill(
        start_color="002060", end_color="002060", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Formatting hyperlink column (index 5)
    link_col_idx = 5
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=link_col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.value = "Apply Here"
            cell.font = Font(color="0563C1", underline="single")

    # Auto-adjust column widths
    for i, col in enumerate(df.columns):
        column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(i + 1)].width = min(
            column_len, 50
        )

    writer.close()
    return filename


# ==========================================
# Google Drive Integration
# ==========================================
def upload_to_drive(drive_service, filename: str, date_str: str) -> tuple:
    try:
        folder_name = f"Job_Scrapes_{date_str}"

        query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false"
        results = (
            drive_service.files()
            .list(q=query, spaces="drive", fields="files(id, webViewLink)")
            .execute()
        )
        items = results.get("files", [])

        if not items:
            folder_metadata = {
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder",
            }
            folder = (
                drive_service.files()
                .create(body=folder_metadata, fields="id, webViewLink")
                .execute()
            )
            folder_id = folder.get("id")
            folder_url = folder.get("webViewLink")
        else:
            folder_id = items[0].get("id")
            folder_url = items[0].get("webViewLink")

        file_metadata = {"name": os.path.basename(filename), "parents": [folder_id]}
        media = MediaFileUpload(filename, resumable=True)
        file = (
            drive_service.files()
            .create(body=file_metadata, media_body=media, fields="id, webViewLink")
            .execute()
        )
        file_id = file.get("id")
        file_url = file.get("webViewLink")

        permission = {"type": "anyone", "role": "reader"}
        drive_service.permissions().create(fileId=file_id, body=permission).execute()

        return folder_id, folder_url, file_id, file_url
    except Exception as e:
        print(f"Failed to upload to Google Drive: {e}")
        print(
            "Note: To fix this, you must share a Google Drive folder with your Service Account email."
        )
        return None, None, None, None


# ==========================================
# Email Notification
# ==========================================
def send_email(
    target_email: str, jobs: List[Dict[str, Any]], filename: str, date_str: str
):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_username = os.getenv("SMTP_USERNAME")
    smtp_password = os.getenv("SMTP_PASSWORD")

    if not smtp_username or not smtp_password:
        print("SMTP credentials missing, skipping email.")
        return False

    msg = EmailMessage()
    msg["Subject"] = f"Job Scraper Report - {date_str}"
    msg["From"] = smtp_username
    msg["To"] = target_email

    total_jobs = len(jobs)
    top_jobs_html = ""
    for job in jobs[:5]:
        link = job.get("Application Link", "#")
        top_jobs_html += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{job.get('Job Title')}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{job.get('Company')}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{job.get('Location')}</td>
            <td style="padding: 8px; border: 1px solid #ddd;"><a href="{link}">Apply</a></td>
        </tr>
        """

    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
        <h2>Automated Job Scraper Report</h2>
        <p>Your scheduled job scrape for <strong>{date_str}</strong> is complete.</p>
        <p><strong>Total Jobs Found:</strong> {total_jobs}</p>
        
        <h3>Top 5 Job Highlights</h3>
        <table style="border-collapse: collapse; width: 100%; max-width: 600px;">
          <thead>
            <tr style="background-color: #002060; color: white;">
              <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Title</th>
              <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Company</th>
              <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Location</th>
              <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Link</th>
            </tr>
          </thead>
          <tbody>
            {top_jobs_html}
          </tbody>
        </table>
        
        <br>
        <p><em>The full list of jobs is attached to this email as an Excel spreadsheet.</em></p>
      </body>
    </html>
    """

    msg.set_content(
        f"Job Scraper Report for {date_str}. Total found: {total_jobs}. See attached spreadsheet."
    )
    msg.add_alternative(html_content, subtype="html")

    # Attach the Excel file
    if os.path.exists(filename):
        with open(filename, "rb") as f:
            file_data = f.read()
            msg.add_attachment(
                file_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(filename),
            )

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False


def should_run_job(db, job_id, scrape_frequency, current_time: datetime) -> bool:
    if scrape_frequency == "Now":
        return True

    runs = (
        db.collection("job_runs")
        .where(filter=FieldFilter("job_id", "==", job_id))
        .order_by("created_at", direction="DESCENDING")
        .limit(1)
        .get()
    )

    if not runs:
        return True

    last_run_doc = runs[0].to_dict()
    created_at = last_run_doc.get("created_at")
    if not created_at:
        return True

    # created_at is typically a DatetimeWithNanoseconds in firestore SDK
    try:
        last_run_time = created_at.replace(tzinfo=None)
    except AttributeError:
        # Fallback if it's not a datetime object
        return True

    hours_diff = (current_time - last_run_time).total_seconds() / 3600

    if scrape_frequency == "Every 4 Hours" and hours_diff >= 3.9:
        return True
    if scrape_frequency == "Every 6 Hours" and hours_diff >= 5.9:
        return True
    if scrape_frequency == "Every 12 Hours" and hours_diff >= 11.9:
        return True
    if scrape_frequency == "Daily" and hours_diff >= 23.9:
        return True

    return False


# ==========================================
# Main Execution Flow
# ==========================================
def main(target_job_id=None):
    start_time = datetime.now()
    date_str = start_time.strftime("%Y-%m-%d")

    print(f"Starting Job Scraper Agent at {start_time}")

    db = initialize_firebase()
    fc_api_key = os.getenv("FIRECRAWL_API_KEY")
    if not fc_api_key:
        print("Missing FIRECRAWL_API_KEY. Exiting.")
        return

    firecrawl = FirecrawlApp(api_key=fc_api_key)

    if target_job_id:
        doc = db.collection("jobs").document(target_job_id).get()
        docs = [doc] if doc.exists else []
    else:
        jobs_ref = db.collection("jobs").where(
            filter=FieldFilter("is_active", "==", True)
        )
        docs = jobs_ref.stream()

    for doc in docs:
        config = doc.to_dict()
        user_id = config.get("user_id")
        job_id = doc.id
        target_email = config.get("target_email")
        scrape_frequency = config.get("scrape_frequency", "Daily")

        if not target_job_id:
            if not should_run_job(db, job_id, scrape_frequency, start_time):
                print(
                    f"Skipping job {job_id} for user {user_id} - not due yet (Freq: {scrape_frequency})."
                )
                continue

        logger = JobLogger()
        logger.log(f"Processing job {job_id} for user {user_id}")

        jobs = scrape_jobs(config, firecrawl, db, user_id, job_id, logger)
        total_found = len(jobs)

        filename = f"Job_Postings_{date_str}_{uuid.uuid4().hex[:6]}.xlsx"
        generate_excel(jobs, filename)

        email_sent = False
        if target_email:
            email_sent = send_email(target_email, jobs, filename, date_str)

        if os.path.exists(filename):
            os.remove(filename)

        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        logger.log(f"Job completed in {execution_time_ms} ms. Total found: {total_found}")

        run_record = {
            "job_id": job_id,
            "user_id": user_id,
            "run_date": date_str,
            "total_found": total_found,
            "drive_folder_id": "",
            "drive_folder_url": "",
            "excel_file_id": "",
            "excel_file_url": "",
            "email_sent": email_sent,
            "status": "SUCCESS" if total_found > 0 else "WARNING",
            "execution_time_ms": execution_time_ms,
            "logs": logger.logs,
            "created_at": firestore.SERVER_TIMESTAMP,
        }

        db.collection("job_runs").add(run_record)
        print(f"Completed run for job {job_id}. Logged to job_runs.")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--job-id", help="Run a specific job by its ID", default=None)
    args = parser.parse_args()

    main(args.job_id)
